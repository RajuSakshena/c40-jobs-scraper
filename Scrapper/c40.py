import os
import time
import json
import re
import pandas as pd
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Alignment

CAREERS_URL = "https://c40.bamboohr.com/careers"
BASE_URL = "https://c40.bamboohr.com"
KEYWORDS_FILE = "keywords.json"

OUTPUT_DIR = "output"
OUTPUT_FILE = "c40_jobs.xlsx"


def load_keywords():
    with open(KEYWORDS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def match_verticals(title, description, keywords):
    text = f"{title} {description}".lower()
    matched = []
    for vertical, words in keywords.items():
        for w in words:
            if re.search(rf"\b{re.escape(w.lower())}\b", text):
                matched.append(vertical)
                break
    return ", ".join(matched) if matched else "N/A"


def scrape_c40_jobs():
    keywords = load_keywords()
    jobs = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        print(f"🔍 Opening listings page: {CAREERS_URL}")
        page.goto(CAREERS_URL, timeout=60000)
        page.wait_for_load_state("networkidle")
        time.sleep(2)

        soup = BeautifulSoup(page.content(), "html.parser")

        job_links = []
        for a in soup.select("a[href^='/careers/']"):
            href = a.get("href")
            title = a.get_text(strip=True)
            if href and title:
                job_links.append((BASE_URL + href, title))

        job_links = list(dict.fromkeys(job_links))
        print(f"✅ Found {len(job_links)} job listings")

        for job_url, fallback_title in job_links:
            print(f"➡️ Visiting job page: {job_url}")
            page.goto(job_url, timeout=60000)
            page.wait_for_load_state("networkidle")
            time.sleep(2)

            job_soup = BeautifulSoup(page.content(), "html.parser")

            # Title
            title_tag = job_soup.find("h3", class_="fabric-oxx0vk-root")
            title = title_tag.get_text(strip=True) if title_tag else fallback_title

            # Description (multiple blocks)
            desc_blocks = job_soup.find_all("div", class_="fabric-95l02p-description")
            description = "\n".join(d.get_text(strip=True) for d in desc_blocks)

            if not description:
                print(f"⚠️ Description missing: {job_url}")
                continue

            matched_vertical = match_verticals(title, description, keywords)

            # Excel-safe hyperlink formula
            excel_safe_title = title.replace('"', "''")

            jobs.append({
                "Title": title,
                "Description": description,
                "Matched_Vertical": matched_vertical,
                "Apply_Link": f'=HYPERLINK("{job_url}", "{excel_safe_title}")'
            })

        browser.close()

    if not jobs:
        print("❌ No jobs extracted.")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    df = pd.DataFrame(
        jobs,
        columns=["Title", "Description", "Matched_Vertical", "Apply_Link"]
    )

    excel_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    df.to_excel(excel_path, index=False, engine="openpyxl")

    # ✅ Excel formatting
    wb = load_workbook(excel_path)
    ws = wb.active

    ws.column_dimensions["A"].width = 55   # Title
    ws.column_dimensions["B"].width = 120  # Description
    ws.column_dimensions["C"].width = 30   # Matched Vertical
    ws.column_dimensions["D"].width = 50   # Apply Link

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(excel_path)

    print(f"✅ Successfully saved {len(df)} jobs to {excel_path}")


if __name__ == "__main__":
    scrape_c40_jobs()
