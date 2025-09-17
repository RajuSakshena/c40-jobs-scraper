import os
import time
import json
import re
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from playwright.sync_api import sync_playwright

CAREERS_URL = "https://www.c40.org/careers/"
KEYWORDS_FILE = "keywords.json"

# === Custom Keywords for "How To Apply" extraction ===
custom_keywords = [
    "Selection Criteria", "Evaluation & Follow-Up", "Application Guidelines", "Eligible Applicants:",
    "Scope of Work:", "Proposal Requirements", "Evaluation Criteria", "Submission Details", "Eligible Entities",
    "How to apply", "Purpose of RFP", "Proposal Guidelines", "Eligibility Criteria", "Application must include:",
    "Eligibility", "Submission of Tender:", "Technical Bid-", "Who Can Apply", "Documents Required", "Expectation:",
    "Eligibility Criterion:", "Submission terms:", "Vendor Qualifications", "To apply",
    "To know about the eligibility criteria:", "The agency's specific responsibilities include –",
    "SELCO Foundation will be responsible for:", "Partner Eligibility Criteria", "Proposal Submission Requirements",
    "Proposal Evaluation Criteria", "Eligibility Criteria for CSOs to be part of the programme:", "Pre-Bid Queries:",
    "Response to Pre-Bid Queries:", "Submission of Bid:", "Applicant Profiles:", "What we like to see in grant applications:",
    "Research that is supported by the SVRI must:", "Successful projects are most often:", "Criteria for funding:",
    "Before you begin to write your proposal, consider that IEF prefers to fund:",
    "As you prepare your budget, these are some items that IEF will not fund:", "Organizational Profile",
    "Selection Process", "Proposal Submission Guidelines", "Terms and Conditions", "Security Deposit:",
    "Facilities and Support Offered under the call for proposal:", "Other Requirements:", "Reporting To:", 
    "Prospective Consultants should demonstrate:", "Term:", "Location:", "Salary:", "Application Process:",
    "Person Specification:", "Position Description:", "Responsibilities:", "Required qualifications:",
    "Modeling, coding, and quantitative tool maintenance", "Specific responsibilities include:"
]

def load_keywords():
    with open(KEYWORDS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def match_verticals(title, description, keywords):
    text_blob = (title + " " + description).lower()
    matched = []
    for vertical, words in keywords.items():
        for word in words:
            if re.search(r"\b" + re.escape(word.lower()) + r"\b", text_blob):
                matched.append(vertical)
                break
    return ", ".join(matched) if matched else "N/A"

def extract_how_to_apply(description):
    """Extract keyword sections with their paragraphs (headings + content)."""
    if not description or not isinstance(description, str):
        return "N/A"

    norm_keywords = [kw.lower().rstrip(":") for kw in custom_keywords]
    matched_sections = []

    # Split into lines
    segments = description.split("\n")
    i = 0
    while i < len(segments):
        seg_clean = segments[i].strip()
        if not seg_clean:
            i += 1
            continue
        seg_lower = seg_clean.lower()

        # If line starts with keyword
        if any(seg_lower.startswith(kw) for kw in norm_keywords):
            section_lines = [f"• {seg_clean}"]
            i += 1
            # Capture following lines until next keyword or empty line
            while i < len(segments):
                next_line = segments[i].strip()
                if not next_line:
                    break
                next_lower = next_line.lower()
                if any(next_lower.startswith(kw) for kw in norm_keywords):
                    break
                section_lines.append(next_line)
                i += 1
            matched_sections.append("\n".join(section_lines))
        else:
            i += 1

    return "\n\n".join(matched_sections) if matched_sections else "N/A"

def scrape_c40_jobs():
    jobs = []
    keywords = load_keywords()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)  # headless for GitHub Actions
        page = browser.new_page()
        
        print(f"🔍 Opening {CAREERS_URL}")
        page.goto(CAREERS_URL, timeout=60000)

        try:
            page.wait_for_selector("a.link-cards-item", timeout=15000)
        except:
            print("⚠️ No job cards loaded within timeout.")
            browser.close()
            return

        soup = BeautifulSoup(page.content(), "html.parser")
        links = soup.select("a.link-cards-item")

        print(f"✅ Found {len(links)} job links")

        for link in links:
            job_url = link["href"].strip()
            job_title = link.find("h3", class_="link-cards-item__heading").get_text(strip=True)

            print(f"➡️ Visiting job page: {job_url}")
            job_page = browser.new_page()
            job_page.goto(job_url, timeout=60000)

            try:
                job_page.wait_for_selector("#descriptionWrapper", timeout=10000)
            except:
                print(f"⚠️ Could not load description for {job_url}")
                job_page.close()
                continue

            detail_soup = BeautifulSoup(job_page.content(), "html.parser")

            # Extract title
            title_tag = detail_soup.find("h3", class_="jss-g20")
            title = title_tag.get_text(strip=True) if title_tag else job_title

            # Extract description
            desc_tag = detail_soup.find("div", id="descriptionWrapper")
            description = desc_tag.get_text(separator="\n", strip=True) if desc_tag else "N/A"

            # Match verticals
            matched_verticals = match_verticals(title, description, keywords)

            # Extract "How to Apply"
            how_to_apply = extract_how_to_apply(description)

            # ✅ Fixed Excel-safe hyperlink string
            excel_safe_title = title.replace('"', "''")
            jobs.append({
                "Title": title,
                "Description": description,
                "How_To_Apply": how_to_apply,
                "Matched_Vertical": matched_verticals,
                "Clickable_Link": f'=HYPERLINK("{job_url}", "{excel_safe_title}")'
            })

            job_page.close()
            time.sleep(1)

        browser.close()

    if not jobs:
        print("⚠️ No jobs found.")
        return

    # Save to Excel in correct column order
    if not os.path.exists("output"):
        os.makedirs("output")

    df = pd.DataFrame(jobs, columns=["Title", "Description", "How_To_Apply", "Matched_Vertical", "Clickable_Link"])

    excel_path = "output/c40_jobs.xlsx"
    df.to_excel(excel_path, index=False, engine="openpyxl")

    # Format Excel
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.column_dimensions["A"].width = 50  # Title
    ws.column_dimensions["B"].width = 120 # Description
    ws.column_dimensions["C"].width = 60  # How_To_Apply
    ws.column_dimensions["D"].width = 30  # Matched_Vertical
    ws.column_dimensions["E"].width = 50  # Clickable Link
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(excel_path)

    print(f"✅ Jobs saved to {excel_path}")

if __name__ == "__main__":
    scrape_c40_jobs()
